import platform
import pandas as pd
from time import sleep
from kivy.properties import NumericProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import Screen
from openpyxl.reader.excel import load_workbook
from app import db_rdmarcas, db_calc_rdmarcas, db_perfumaria, db_calc_perfumaria, db_dermo, db_calc_dermo, \
    popup_Confirmacao_Exclusao, popupError
from app import click_button, back_button

# Verifica se o usuário está usando Windows
if platform.system() == "Windows":
    sistema_windows = True
    font_column = 18
    font_row = 16
    font_button = 35
    font_text = 35
    font_text_menu = 48
    font_title = 60

else:
    sistema_windows = False
    font_column = 20
    font_row = 19
    font_button = 55
    font_text = 40
    font_text_menu = 60
    font_title = 80


class LimparDados(Screen):
    """
    Opção para limpar dos dados, porém se faz necessário escolher as
    lista que deseja limpar os dados,
    poderá escolher entre: RD Marcas, Perfumaria, dermo ou todas ao mesmo tempo.
    """

    font_column = NumericProperty(font_column)
    font_row = NumericProperty(font_row)
    font_button = NumericProperty(font_button)
    font_text = NumericProperty(font_text)
    font_text_menu = NumericProperty(font_text_menu)
    font_title = NumericProperty(font_title)

    def pressButton(self):
        click_button.play()

    def pressBackButton(self):
        back_button.play()

    def apagarLista_popup(self):
        """
        --> Função que mostra um Popup de confirmação antes de prosseguir com a exclusão das listas.
        """

        content = BoxLayout(orientation='vertical', padding=10)
        label = Label(text='Confirma a exclusão das Listas?')

        close_button = Button(text='Cancelar', size_hint=(1.0, 0.2))
        confirm_button = Button(text='Confirmar', size_hint=(1.0, 0.2))

        content.add_widget(label)
        content.add_widget(close_button)
        content.add_widget(confirm_button)

        popup = Popup(title='Aviso', content=content, size_hint=(0.7, 0.5))

        close_button.bind(on_release=popup.dismiss)

        confirm_button.bind(on_release=lambda btn: self.apagarLista())
        confirm_button.bind(on_release=popup.dismiss)
        popup.open()

    def apagarLista(self):
        """
        --> Função que apaga todas as listas após confirmar no popup o procedimento.
        """

        try:
            # Exclusão RD MARCAS
            # Carrega o arquivo
            lista = db_rdmarcas
            calculo = db_calc_rdmarcas
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_RDMarcas = pd.read_excel(lista)
            calc_lista_RDMarcas = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_RDMarcas.to_excel(db_rdmarcas, index=False)
            calc_lista_RDMarcas.to_excel(db_calc_rdmarcas, index=False)

            # Exclusão PERFUMARIA
            # Carrega o arquivo
            lista = db_perfumaria
            calculo = db_calc_perfumaria
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Perfumaria = pd.read_excel(lista)
            calc_lista_Perfumaria = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Perfumaria.to_excel(db_perfumaria, index=False)
            calc_lista_Perfumaria.to_excel(db_calc_perfumaria, index=False)

            # Exclusão dermo
            # Carrega o arquivo
            lista = db_dermo
            calculo = db_calc_dermo
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Dermo = pd.read_excel(lista)
            calc_lista_Dermo = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Dermo.to_excel(db_dermo, index=False)
            calc_lista_Dermo.to_excel(db_calc_dermo, index=False)

            # POPUP DE FINALIZAÇÃO
            sleep(0.3)
            popup_Confirmacao_Exclusao()

        except Exception as error:
            print(error)
            popupError()


class LimparTodasAsListas(Screen):
    """
    """
    pass
