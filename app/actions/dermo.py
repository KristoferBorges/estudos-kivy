import platform
import pandas as pd
from time import sleep
from kivy.core.window import Window
from kivy.properties import NumericProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from openpyxl.reader.excel import load_workbook
from app import popup_Confirmacao_Exclusao, popupError, dateVerification, abatimento, formataLista, db_dermo, \
    db_calc_dermo
from app import click_button, back_button

# Variável para testar inserções de dados
teste = False

# Verifica se o usuário está usando Windows
if platform.system() == "Windows":
    sistema_windows = True
    font_column = 18
    font_row = 16
    font_button = 35
    font_text = 35
    font_text_menu = 48
    font_title = 60

else:
    sistema_windows = False
    font_column = 20
    font_row = 19
    font_button = 55
    font_text = 40
    font_text_menu = 60
    font_title = 80


class RegistrosDermo(Screen):
    """
    Opção do menu principal após clicar na opção de registros (Dermo).
    """

    font_column = NumericProperty(font_column)
    font_row = NumericProperty(font_row)
    font_button = NumericProperty(font_button)
    font_text = NumericProperty(font_text)
    font_text_menu = NumericProperty(font_text_menu)
    font_title = NumericProperty(font_title)

    def pressButton(self):
        click_button.play()

    def pressBackButton(self):
        back_button.play()

    def pega_input_dermo(self):
        """
        --> Função para pegar os dados inseridos na opção 'REGISTROS' -> 'DERMO'.
        :return: Retorna os dados devidamente formatados .
        """
        try:
            data = self.ids.data_input.text
            data = dateVerification(data)
            metaDia = float(self.ids.meta_input.text.replace('-', '').strip())
            vendaDia = float(self.ids.venda_input.text.replace('-', '').strip())
            pecaDia = int(self.ids.peca_input.text.replace('-', '').strip())

            # Exibição no terminal
            pd.set_option('display.max_columns', None)  # Exibe todas as colunas
            pd.set_option('display.max_rows', None)  # Exibe todas as linhas
            pd.set_option('display.width', 1000)  # Largura máxima da exibição

            # Carrega o arquivo
            df_lista_Dermo = pd.read_excel(db_dermo)
            calc_lista_Dermo = pd.read_excel(db_calc_dermo)

            novoCalc = {
                'Meta': f'{metaDia}',
                'Venda': f'{vendaDia}',
                'Pecas': f'{pecaDia}',
            }

            # Insere o dado na lista
            calc_lista_Dermo.loc[len(calc_lista_Dermo)] = novoCalc

            metaAC = calc_lista_Dermo['Meta'].astype(float).sum()
            vendaAC = calc_lista_Dermo['Venda'].astype(float).sum()
            pecaAC = calc_lista_Dermo['Pecas'].astype(float).sum()

            if vendaAC < metaAC:
                sobras = (metaAC - vendaAC)
            elif metaAC < vendaAC:
                sobras = (vendaAC - metaAC)
            else:
                sobras = 0

            if vendaAC != 0 and metaAC != 0:
                porcentagem = (vendaAC / metaAC) * 100
            else:
                porcentagem = 'Error'

            # Análise alcance de metas
            devedor = abatimento(metaAC, vendaAC)

            # Cria uma linha de inserção de dados
            novoDado = {
                'Data': f'{data}',
                'Meta': f'{metaDia:.2f}',
                'Meta.AC': f'{metaAC:.2f}',
                'Venda': f'{vendaDia:.2f}',
                'Venda.AC': f'{vendaAC:.2f}',
                'Pecas.AC': f'{pecaAC}',
                'Sobras': f'{sobras:.2f}',
                'P': f'{porcentagem:.2f}'
            }

            # Insere o dado na lista
            df_lista_Dermo.loc[len(df_lista_Dermo)] = novoDado
            print(df_lista_Dermo)
            print(f'Meta: {metaAC}')
            print(f'Venda: {vendaAC}')
            print(f'Peças: {pecaAC}')

            # Salva o arquivo com o novo dado
            df_lista_Dermo.to_excel(db_dermo, index=False)
            calc_lista_Dermo.to_excel(db_calc_dermo, index=False)

            # Limpa os dados anteriormente informados (Somente teste = False)
            if not teste:
                self.ids.data_input.text = ""
                self.ids.meta_input.text = ""
                self.ids.venda_input.text = ""
                self.ids.peca_input.text = ""

                # Baseado na variável (devedor) o sistema passará a situação da meta/vendas no popup
                if devedor == '-':
                    situacao = "Metas não atingidas!"
                else:
                    situacao = "Metas atingitas!"

                # Popup de resumo
                content = BoxLayout(orientation='vertical', padding=10)
                label = Label(text=f'Resumo Acumulado (DERMO)\n\n'
                                   f'Meta: R$ {metaAC:.2f}\nVendas: R$ {vendaAC:.2f}\nPeças: {pecaAC} Un\n'
                                   f'Sobras: {devedor}R$ {sobras:.2f}\nSituação: {situacao}\n')
                close_button = Button(text='Fechar', size_hint=(1.0, 0.2))

                content.add_widget(label)
                content.add_widget(close_button)

                popup = Popup(title='Dados armazenados com Sucesso!', content=content, size_hint=(0.7, 0.5))
                close_button.bind(on_release=popup.dismiss)
                popup.open()

        except ValueError as error:
            print(error)
            content = BoxLayout(orientation='vertical', padding=10)
            label = Label(text='Campos não preenchidos!')
            close_button = Button(text='Fechar', size_hint=(1.0, 0.2))

            content.add_widget(label)
            content.add_widget(close_button)

            popup = Popup(title='Aviso', content=content, size_hint=(0.5, 0.3))
            close_button.bind(on_release=popup.dismiss, on_press=lambda btn: self.pressButton())
            popup.open()

        except Exception as error:
            print(error)
            popupError()


class LimparDermo(Screen):
    """
    """

    def __init__(self, **kw):
        super().__init__()
        self.tipo_busca = None
        self.index_value = None
        self.max_lines = None
        self.df_lista_Dermo = None
        self.calc_lista_Dermo = None

    font_column = NumericProperty(font_column)
    font_row = NumericProperty(font_row)
    font_button = NumericProperty(font_button)
    font_text = NumericProperty(font_text)
    font_text_menu = NumericProperty(font_text_menu)
    font_title = NumericProperty(font_title)

    def pressButton(self):
        click_button.play()

    def pressBackButton(self):
        back_button.play()


    def apagarLista_popup_Dermo(self):
        """
        --> Função que mostra um Popup de confirmação antes de prosseguir com a exclusão das listas.
        """

        content = BoxLayout(orientation='vertical', padding=10)
        label = Label(text='Confirma a exclusão da Lista\nDermo?')

        close_button = Button(text='Cancelar', size_hint=(1.0, 0.2))
        confirm_button = Button(text='Confirmar', size_hint=(1.0, 0.2))

        content.add_widget(label)
        content.add_widget(close_button)
        content.add_widget(confirm_button)

        popup = Popup(title='Aviso', content=content, size_hint=(0.7, 0.5))

        close_button.bind(on_release=popup.dismiss, on_press=lambda btn: self.pressButton())

        confirm_button.bind(on_release=lambda btn: self.apagarLista_Dermo())
        confirm_button.bind(on_release=popup.dismiss, on_press=lambda btn: self.pressButton())
        popup.open()

    def apagarLista_Dermo(self):
        """
        --> Função que apaga a lista Dermo após confirmar no popup o procedimento.
        """

        try:
            # Exclusão Dermo
            # Carrega o arquivo
            lista = db_dermo
            calculo = db_calc_dermo
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Dermo = pd.read_excel(lista)
            calc_lista_Dermo = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Dermo.to_excel(db_dermo, index=False)
            calc_lista_Dermo.to_excel(db_calc_dermo, index=False)

            # POPUP DE FINALIZAÇÃO
            sleep(0.3)
            popup_Confirmacao_Exclusao()

        except Exception as error:
            print(error)
            popupError()

    def buscarPesquisa(self):
        try:
            # Carregar o arquivo (O arquivo de calculo e o arquivo da lista de visualização)
            global linha_filtrada
            self.calc_lista_Dermo = pd.read_excel(db_calc_dermo)
            self.df_lista_Dermo = pd.read_excel(db_dermo)

            # Verifica a quantidade máxima de linhas dentro do arquivo
            self.max_lines = len(
                self.calc_lista_Dermo)  # Os dois aquivos importados tem a mesma quantidade de linha

            # Localiza a linha com base no input do usuário
            busca = self.ids.research_input.text

            # Limpa o resultado das buscas anteriores
            self.ids.resultado_linha.text = ''

            if len(str(busca)) > 4:
                linha_filtrada = self.df_lista_Dermo[self.df_lista_Dermo['Data'] == busca]
                self.ids.busca_resultado.text = f'Informação encontrada!'
                self.ids.resultado_linha.text = f'{linha_filtrada}'
                self.tipo_busca = 'data'

            elif int(busca.isnumeric()) and int(busca) - 2 <= self.max_lines - 1 and int(busca) >= 2:  # busca - 2
                busca = int(busca) - 2
                linha_filtrada = self.df_lista_Dermo[self.df_lista_Dermo.index == busca]
                self.ids.busca_resultado.text = f'Informação encontrada!'
                self.ids.resultado_linha.text = f'{linha_filtrada}'
                self.tipo_busca = 'index'

            else:
                self.ids.busca_resultado.text = 'Informação não Localizada!'
                self.ids.finalizar_alteracao.text = ''
                self.ids.resultado_linha.text = ''
                self.tipo_busca = None

            self.index_value = linha_filtrada.index[0]
            print(self.index_value)

            return self.calc_lista_Dermo, self.df_lista_Dermo, self.max_lines, self.index_value

        except Exception as error:
            print(f'Houve um erro - {error}')
            self.ids.busca_resultado.text = 'Informação não Localizada!'

    def executarAlteracao(self):
        try:
            # Define a quantidade de repetições iniciais
            qnt = 1

            # Cálculo dos dados
            data = self.ids.data_input.text
            metaDia = float(self.ids.meta_input.text.replace('-', '').strip())
            vendaDia = float(self.ids.venda_input.text.replace('-', '').strip())
            pecaDia = int(self.ids.peca_input.text.replace('-', '').strip())
            calc_lista_Dermo = self.calc_lista_Dermo
            for _ in calc_lista_Dermo.iterrows():
                if qnt == 1:
                    qnt = qnt + 1
                    # Insere os valores (MetaDia/VendaDia),
                    # logo em seguida é feito o cálculo já pegando o valor alterado
                    self.calc_lista_Dermo.loc[self.index_value] = [metaDia, vendaDia, pecaDia]

                    metaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Meta'].astype(float).sum()
                    vendaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Venda'].astype(float).sum()
                    pecaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Pecas'].astype(int).sum()
                    if vendaAC < metaAC:
                        sobras = (metaAC - vendaAC)
                    elif metaAC < vendaAC:
                        sobras = (vendaAC - metaAC)
                    else:
                        sobras = 0

                    if vendaAC != 0 and metaAC != 0:
                        porcentagem = (vendaAC / metaAC) * 100
                    else:
                        porcentagem = 'Error'

                    # Input de dados
                    novoDado = {
                        'Data': f'{data}',
                        'Meta': f'{metaDia:.2f}',
                        'Meta.AC': f'{metaAC:.2f}',
                        'Venda': f'{vendaDia:.2f}',
                        'Venda.AC': f'{vendaAC:.2f}',
                        'Pecas.AC': f'{pecaAC}',
                        'Sobras': f'{sobras:.2f}',
                        'P': f'{porcentagem:.2f}'
                    }

                    # Modifica os valores da linha (MetaDia/VendaDia) | Modifica os dados da
                    # linha por completo com os devidos cálculos
                    self.df_lista_Dermo.loc[self.index_value] = novoDado

                    # Salva o arquivo
                    self.df_lista_Dermo.to_excel(db_dermo, index=False)
                    self.calc_lista_Dermo.to_excel(db_calc_dermo, index=False)

                elif self.max_lines >= self.index_value and self.index_value < self.max_lines:
                    self.index_value = self.index_value + 1

                    metaDia = self.calc_lista_Dermo.at[self.index_value, 'Meta']
                    vendaDia = self.calc_lista_Dermo.at[self.index_value, 'Venda']
                    pecaDia = self.calc_lista_Dermo.at[self.index_value, 'Pecas']

                    data_atualizada = self.df_lista_Dermo.at[self.index_value, 'Data']
                    self.calc_lista_Dermo.loc[self.index_value] = [metaDia, vendaDia, pecaDia]

                    metaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Meta'].astype(float).sum()
                    vendaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Venda'].astype(float).sum()
                    pecaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Pecas'].astype(int).sum()

                    if vendaAC < metaAC:
                        sobras = (metaAC - vendaAC)
                    elif metaAC < vendaAC:
                        sobras = (vendaAC - metaAC)
                    else:
                        sobras = 0

                    if vendaAC != 0 and metaAC != 0:
                        porcentagem = (vendaAC / metaAC) * 100
                    else:
                        porcentagem = 'Error'

                    # Input de dados
                    novoDado = {
                        'Data': f'{data_atualizada}',
                        'Meta': f'{metaDia:.2f}',
                        'Meta.AC': f'{metaAC:.2f}',
                        'Venda': f'{vendaDia:.2f}',
                        'Venda.AC': f'{vendaAC:.2f}',
                        'Pecas.AC': f'{pecaAC}',
                        'Sobras': f'{sobras:.2f}',
                        'P': f'{porcentagem:.2f}'
                    }

                    self.df_lista_Dermo.loc[self.index_value] = novoDado

                    # Salva o arquivo
                    self.df_lista_Dermo.to_excel(db_dermo, index=False)
                    self.calc_lista_Dermo.to_excel(db_calc_dermo, index=False)

                    # Texto do label de confimação após alterações
                    self.ids.finalizar_alteracao.text = 'Alterações realizadas'

                    # Limpa os dados inseridos e coloca a data de alteração no campo de pesquisa
                    if self.tipo_busca == 'data':
                        self.ids.research_input.text = f'{data}'

                    # Limpa os inputs preenchidos
                    self.ids.data_input.text = ''
                    self.ids.meta_input.text = ''
                    self.ids.venda_input.text = ''
                    self.ids.peca_input.text = ''

        except Exception as error:
            print(f'Houve um erro - {error}')
            if self.tipo_busca == 'data':
                self.ids.finalizar_alteracao.text = 'Faça outra Busca\n  Para atualizar'
            elif self.tipo_busca == 'index':
                self.buscarPesquisa()
            else:
                self.ids.finalizar_alteracao.text = 'Faça outra Busca\n  Para atualizar'


class ConsultaDermo(Screen):
    """
    --> Opção que consulta a lista Dermo.
    """

    font_column = NumericProperty(font_column)
    font_row = NumericProperty(font_row)
    font_button = NumericProperty(font_button)
    font_text = NumericProperty(font_text)
    font_text_menu = NumericProperty(font_text_menu)
    font_title = NumericProperty(font_title)

    def pressButton(self):
        click_button.play()

    def pressBackButton(self):
        back_button.play()

    def mostrarLista(self):
        if sistema_windows:
            Window.size = (1130, 810)
        df_lista_Dermo = pd.read_excel(db_dermo)
        df_lista_Dermo = formataLista(lista=df_lista_Dermo, button='DERMO')

        # Cria um layout para a tabela
        layout = BoxLayout(orientation='vertical', spacing=10)

        # Cria um ScrollView para permitir a rolagem vertical
        scrollview = ScrollView()

        # Cria um BoxLayout para conter os labels do cabeçalho
        header_layout = BoxLayout(orientation='horizontal', spacing=10, size_hint_y=None, height=30)

        # Obtém as colunas da tabela
        columns = df_lista_Dermo.columns

        # Cria labels para os nomes das colunas e adiciona ao header_layout
        for column in columns:
            header_label = Label(font_size=font_column, text=column, size_hint_x=None, width=116, halign='left')
            header_layout.add_widget(header_label)

        # Adiciona o header_layout ao layout principal
        layout.add_widget(header_layout)

        # Cria um BoxLayout para conter as labels da tabela
        table_layout = BoxLayout(orientation='vertical', spacing=5, padding=10, size_hint_y=None)

        # Itera pelas linhas e cria labels para os valores
        for index, row in df_lista_Dermo.iterrows():
            # Cria um novo layout para cada linha da tabela
            row_layout = BoxLayout(orientation='horizontal', spacing=10, size_hint_y=None, height=30)

            values = [str(row[column]) for column in columns]
            for value in values:
                # Adiciona cada valor em uma label separada dentro do novo layout
                row_labels = Label(font_size=font_row, text=value, size_hint_x=None, width=114, halign='left')
                row_layout.add_widget(row_labels)

            table_layout.add_widget(row_layout)

        # Define a altura do table_layout com base no número de linhas
        table_layout.height = len(df_lista_Dermo) * 35

        scrollview.add_widget(table_layout)
        layout.add_widget(scrollview)

        # Cria um Popup e define o seu conteúdo como o layout da tabela
        popup = Popup(title='Tabela Dermo', content=layout, size_hint=(0.9, 0.8))
        popup.open()
